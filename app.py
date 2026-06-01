#!/usr/bin/env python3
"""
app.py — Collegedunia Article Ranking Pipeline — Web UI  (v2)
==============================================================
Start:
    pip3 install flask reportlab --break-system-packages
    python3 app.py

Then open:  http://localhost:5050
"""

import os, sys, json, csv, time, threading, subprocess, smtplib, logging, shutil, signal
from datetime import datetime, date, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text      import MIMEText
from email.mime.base      import MIMEBase
from email                import encoders
from io                   import BytesIO, StringIO

from flask import (
    Flask, render_template, Response, request,
    jsonify, stream_with_context, send_file
)

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    handlers=[
        logging.FileHandler("pipeline_ui.log"),
        logging.StreamHandler(sys.stdout),
    ]
)
log = logging.getLogger("ui")

# ── Python executable (safe for PythonAnywhere where sys.executable = uwsgi) ─
def _find_python() -> str:
    """Return a Python interpreter that can import gspread.

    On PythonAnywhere the web app runs under uwsgi, so sys.executable is
    the uwsgi binary, not Python.  This function:
      1. Tries sys.executable first (works on local dev / GitHub Actions).
      2. Searches the user's virtualenvs (PythonAnywhere virtualenv setup).
      3. Tries known absolute Python paths on PythonAnywhere.
      4. Falls back to PATH search.
    Each candidate is tested with a quick `import gspread` to confirm it has
    the right packages installed.
    """
    def _test(path: str) -> bool:
        """Return True if 'path' is a Python executable that has gspread."""
        try:
            r = subprocess.run(
                [path, "-c", "import gspread"],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                timeout=10
            )
            return r.returncode == 0
        except Exception:
            return False

    def _is_python(path: str) -> bool:
        return bool(path) and "python" in os.path.basename(path).lower()

    candidates = []

    # 1. sys.executable (real Python on local/CI; uwsgi on PythonAnywhere)
    exe = sys.executable or ""
    if _is_python(exe):
        candidates.append(exe)

    # 2. Virtualenvs (PythonAnywhere web apps often use a venv)
    home = os.path.expanduser("~")
    for venv_root in [
        os.path.join(home, ".virtualenvs"),
        os.path.join(home, "venv"),
        os.path.join(home, "envs"),
    ]:
        if os.path.isdir(venv_root):
            for venv_name in sorted(os.listdir(venv_root)):
                for py in ("python3.11", "python3", "python"):
                    p = os.path.join(venv_root, venv_name, "bin", py)
                    if os.path.isfile(p) and p not in candidates:
                        candidates.append(p)

    # 3. Known absolute paths on PythonAnywhere (no PATH lookup needed)
    for p in (
        "/usr/bin/python3.11",
        "/usr/bin/python3.10",
        "/usr/bin/python3.9",
        "/usr/bin/python3",
        "/usr/local/bin/python3.11",
        "/usr/local/bin/python3",
    ):
        if p not in candidates:
            candidates.append(p)

    # 4. PATH-based search
    for name in ("python3.11", "python3.10", "python3.9", "python3", "python"):
        found = shutil.which(name)
        if found and found not in candidates:
            candidates.append(found)

    # Test each candidate; return the first that has gspread
    for candidate in candidates:
        if not os.path.isfile(candidate):
            continue
        if _test(candidate):
            log.info(f"PYTHON_EXE resolved to: {candidate}")
            return candidate

    # Last resort: return sys.executable even if gspread test failed
    fallback = exe if _is_python(exe) else (shutil.which("python3") or "python3")
    log.warning(f"Could not find Python with gspread — falling back to: {fallback}")
    return fallback

PYTHON_EXE = _find_python()

# ── Paths ────────────────────────────────────────────────────────────────────
BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
STATE_FILE      = os.path.join(BASE_DIR, "pipeline_state.json")
UI_CONFIG_FILE  = os.path.join(BASE_DIR, "ui_config.json")
PIPELINE_LOG    = os.path.join(BASE_DIR, "pipeline_ui.log")

app = Flask(__name__)

# ── Runtime state ────────────────────────────────────────────────────────────
_lock            = threading.Lock()
_running         = {"pipeline": False, "quickrun": False}
_procs           = {"pipeline": None,  "quickrun": None}   # live subprocess handles
_next_auto_run   = ""          # ISO timestamp of next midnight auto-run
_midnight_timer  = None        # threading.Timer handle


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS — State / Config
# ══════════════════════════════════════════════════════════════════════════════

def load_state() -> dict:
    try:
        with open(STATE_FILE) as f:
            return json.load(f)
    except Exception:
        return {}


def load_ui_config() -> dict:
    defaults = {
        "smtp_host": "", "smtp_port": 587,
        "smtp_user": "", "smtp_password": "",
        "smtp_from": "", "recipients": []
    }
    if os.path.exists(UI_CONFIG_FILE):
        try:
            with open(UI_CONFIG_FILE) as f:
                data = json.load(f)
            defaults.update(data)
        except Exception:
            pass
    return defaults


def save_ui_config(cfg: dict):
    with open(UI_CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS — Google Sheets (read-only)
# ══════════════════════════════════════════════════════════════════════════════

def get_sheet_data(sheet_name: str, max_rows: int = 500):
    try:
        import gspread
        from google.oauth2.credentials import Credentials
        token_file = os.path.join(BASE_DIR, "token.json")
        if not os.path.exists(token_file):
            return []
        SCOPES = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive.file",
        ]
        creds = Credentials.from_authorized_user_file(token_file, SCOPES)
        gc    = gspread.authorize(creds)
        state = load_state()
        sh    = gc.open_by_key(state["spreadsheet_id"])
        ws    = sh.worksheet(sheet_name)
        rows  = ws.get_all_values()
        return rows[:max_rows]
    except Exception as e:
        log.warning(f"get_sheet_data({sheet_name}): {e}")
        return []


# ══════════════════════════════════════════════════════════════════════════════
#  SSE STREAMING
# ══════════════════════════════════════════════════════════════════════════════

def sse_event(data: str, event: str = "message") -> str:
    out = ""
    for line in data.splitlines():
        out += f"data: {line}\n"
    out += f"event: {event}\n\n"
    return out


def stream_subprocess(cmd: list, stdin_data: str = None, cwd: str = None,
                       proc_key: str = None):
    """Stream a subprocess as SSE events.  If proc_key is given ('pipeline' or
    'quickrun') the Popen handle is stored in _procs so it can be terminated."""
    env = os.environ.copy()
    env["PYTHONUNBUFFERED"] = "1"

    proc = subprocess.Popen(
        cmd,
        stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
        stdin=subprocess.PIPE if stdin_data else None,
        cwd=cwd or BASE_DIR, env=env, text=True, bufsize=1,
        preexec_fn=os.setsid,   # new process group → lets us kill the whole tree
    )

    if proc_key:
        _procs[proc_key] = proc

    if stdin_data:
        def feed():
            try:
                proc.stdin.write(stdin_data)
                proc.stdin.close()
            except Exception:
                pass
        threading.Thread(target=feed, daemon=True).start()

    for line in proc.stdout:
        yield sse_event(line.rstrip())

    proc.wait()
    if proc_key:
        _procs[proc_key] = None

    if proc.returncode in (0, 2):
        yield sse_event("", event="done")
    elif proc.returncode in (-15, -9):    # SIGTERM / SIGKILL — user terminated
        yield sse_event("⛔ Terminated by user.", event="done")
    else:
        yield sse_event(f"Process exited with code {proc.returncode}", event="error")


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Pages
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return render_template("index.html")


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Debug
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/debug")
def api_debug():
    """Diagnostic endpoint — shows which Python will be used for subprocesses."""
    gspread_ok = False
    try:
        import gspread as _gs  # noqa
        gspread_ok = True
    except ImportError:
        pass

    return jsonify({
        "sys_executable": sys.executable,
        "sys_version":    sys.version,
        "PYTHON_EXE":     PYTHON_EXE,
        "gspread_importable_in_webprocess": gspread_ok,
        "BASE_DIR":       BASE_DIR,
    })


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Status
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/status")
def api_status():
    state           = load_state()
    CYCLE_DAYS      = 15
    next_batch_date = ""
    next_batch_days = None
    phase1_last_run = state.get("phase1_last_run", "")
    if phase1_last_run and phase1_last_run != "Never":
        try:
            last_run_d      = date.fromisoformat(phase1_last_run)
            next_d          = last_run_d + timedelta(days=CYCLE_DAYS)
            next_batch_date = next_d.isoformat()
            next_batch_days = (next_d - date.today()).days
        except Exception:
            pass

    return jsonify({
        "run_number":        state.get("run_number", "—"),
        "phase1_last_run":   state.get("phase1_last_run", "Never"),
        "pipeline_locked":   "pipeline_lock" in state,
        "lock_since":        state.get("pipeline_lock", ""),
        "pipeline_running":  _running["pipeline"],
        "quickrun_running":  _running["quickrun"],
        "spreadsheet_id":    state.get("spreadsheet_id", ""),
        "next_auto_run":     _next_auto_run,
        "next_batch_date":   next_batch_date,
        "next_batch_days":   next_batch_days,
    })


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Run Pipeline (SSE)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/run-pipeline")
def api_run_pipeline():
    with _lock:
        if _running["pipeline"]:
            return jsonify({"error": "Pipeline is already running."}), 409
        if _running["quickrun"]:
            return jsonify({"error": "Quick Run is in progress — wait for it to finish."}), 409
        _running["pipeline"] = True

    def generate():
        try:
            yield sse_event("▶  Starting full pipeline …")
            cmd = [PYTHON_EXE, os.path.join(BASE_DIR, "run_pipeline.py")]
            yield from stream_subprocess(cmd, proc_key="pipeline")
        finally:
            _running["pipeline"] = False

    return Response(stream_with_context(generate()), mimetype="text/event-stream",
                    headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"})


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Quick Run (SSE via POST + fetch stream)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/quick-run", methods=["POST"])
def api_quick_run():
    body  = request.get_json(force=True, silent=True) or {}
    mode  = body.get("mode", "").strip()
    value = body.get("value", "").strip()

    if mode not in ("batch", "college"):
        return jsonify({"error": "mode must be 'batch' or 'college'"}), 400
    if not value:
        return jsonify({"error": "value is required"}), 400

    with _lock:
        if _running["pipeline"]:
            return jsonify({"error": "Main pipeline is running — wait for it to finish."}), 409
        if _running["quickrun"]:
            return jsonify({"error": "A Quick Run is already in progress."}), 409
        _running["quickrun"] = True

    mode_num  = "1" if mode == "batch" else "2"
    stdin_str = f"{mode_num}\n{value}\ny\n"

    def generate():
        try:
            yield sse_event(f"▶  Starting Quick Run ({mode}: {value}) …")
            cmd = [PYTHON_EXE, os.path.join(BASE_DIR, "quick_run.py")]
            yield from stream_subprocess(cmd, stdin_data=stdin_str, proc_key="quickrun")
        finally:
            _running["quickrun"] = False

    return Response(stream_with_context(generate()), mimetype="text/event-stream",
                    headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"})


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Clear Lock
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/clear-lock", methods=["POST"])
def api_clear_lock():
    try:
        state = load_state()
        if "pipeline_lock" not in state:
            return jsonify({"message": "No active lock found."})
        state.pop("pipeline_lock")
        with open(STATE_FILE, "w") as f:
            json.dump(state, f, indent=2)
        log.info("Pipeline lock cleared via UI.")
        return jsonify({"message": "Pipeline lock cleared successfully."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Terminate
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/terminate", methods=["POST"])
def api_terminate():
    """Kill the currently running pipeline or quick run subprocess (and all children)."""
    body   = request.get_json(force=True, silent=True) or {}
    target = body.get("target", "")   # "pipeline" or "quickrun"

    if target not in ("pipeline", "quickrun"):
        return jsonify({"error": "target must be 'pipeline' or 'quickrun'"}), 400

    proc = _procs.get(target)
    if proc is None or proc.poll() is not None:
        return jsonify({"message": "No running process found."})

    try:
        # Kill the entire process group — catches child processes (e.g. Apps Script waits)
        try:
            pgid = os.getpgid(proc.pid)
            os.killpg(pgid, signal.SIGKILL)   # SIGKILL cannot be caught or ignored
            log.info(f"SIGKILL sent to process group {pgid} ({target})")
        except ProcessLookupError:
            pass   # process already exited
        except Exception:
            proc.kill()   # fallback: kill just the direct process
            log.info(f"proc.kill() fallback for: {target}")

        return jsonify({"message": f"⛔ {target.capitalize()} terminated. "
                                    "Data generated so far is preserved."})
    except Exception as e:
        log.warning(f"Terminate error ({target}): {e}")
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Sheet GIDs  (so the UI can embed the right tab via iframe)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/sheet-gids")
def api_sheet_gids():
    """Return the numeric GID for the Final and Quick Run Final sheets."""
    try:
        import gspread
        from google.oauth2.credentials import Credentials
        token_file = os.path.join(BASE_DIR, "token.json")
        if not os.path.exists(token_file):
            return jsonify({"error": "token.json not found"}), 500
        SCOPES = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive.file",
        ]
        creds = Credentials.from_authorized_user_file(token_file, SCOPES)
        gc    = gspread.authorize(creds)
        state = load_state()
        sh    = gc.open_by_key(state["spreadsheet_id"])
        gids  = {}
        for ws in sh.worksheets():
            if ws.title in ("Final", "Quick Run Final"):
                gids[ws.title] = ws.id
        return jsonify({
            "spreadsheet_id": state.get("spreadsheet_id", ""),
            "gids": gids,
        })
    except Exception as e:
        log.warning(f"api_sheet_gids: {e}")
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Rankings Data  (source = pipeline | quickrun)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/data/rankings")
def api_rankings():
    source     = request.args.get("source", "pipeline")
    sheet_name = "Final" if source == "pipeline" else "Quick Run Final"
    rows = get_sheet_data(sheet_name, max_rows=5000)
    if not rows:
        return jsonify({"headers": [], "rows": [], "source": source, "total": 0})
    return jsonify({
        "headers": rows[0],
        "rows":    rows[1:],
        "source":  source,
        "total":   len(rows) - 1,
    })


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Download CSV
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/download/csv")
def api_download_csv():
    source     = request.args.get("source", "pipeline")
    sheet_name = "Final" if source == "pipeline" else "Quick Run Final"
    rows = get_sheet_data(sheet_name, max_rows=5000)
    if not rows:
        return jsonify({"error": "No data found in sheet."}), 404

    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerows(rows)
    buf.seek(0)

    fname = f"rankings_{source}_{date.today().isoformat()}.csv"
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Download Rankings PDF  (table of all ranking rows)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/download/rankings-pdf")
def api_download_rankings_pdf():
    source     = request.args.get("source", "pipeline")
    sheet_name = "Final" if source == "pipeline" else "Quick Run Final"
    rows = get_sheet_data(sheet_name, max_rows=5000)
    if not rows or len(rows) < 2:
        return jsonify({"error": "No data found in sheet."}), 404
    try:
        buf   = _generate_rankings_pdf(rows, source)
        fname = f"rankings_{source}_{date.today().isoformat()}.pdf"
        return send_file(buf, mimetype="application/pdf",
                         as_attachment=True, download_name=fname)
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        log.exception("Rankings PDF error")
        return jsonify({"error": str(e)}), 500


def _truncate_to_width(text, max_width, font_name="Helvetica", font_size=7.5):
    """Clip text to a single line that fits `max_width` points, adding an
    ellipsis when truncated. Used for PDF cells only so each cell stays inside
    its column borders (never wraps to a taller row, never overflows)."""
    from reportlab.pdfbase.pdfmetrics import stringWidth
    text = str(text)
    if max_width <= 0:
        return ""
    if stringWidth(text, font_name, font_size) <= max_width:
        return text
    ell = "…"
    while text and stringWidth(text + ell, font_name, font_size) > max_width:
        text = text[:-1]
    return (text + ell) if text else ell


def _generate_rankings_pdf(rows: list, source: str) -> BytesIO:
    try:
        from reportlab.lib.pagesizes  import A4, landscape
        from reportlab.lib.styles     import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units      import cm
        from reportlab.lib            import colors
        from reportlab.platypus       import (
            SimpleDocTemplate, Paragraph, Spacer,
            Table, TableStyle, HRFlowable
        )
    except ImportError:
        raise RuntimeError("reportlab not installed — run: pip3 install reportlab --break-system-packages")

    def _ep(s):
        """Escape text for use inside a reportlab Paragraph."""
        return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=0.8*cm, rightMargin=0.8*cm,
                            topMargin=1.0*cm, bottomMargin=1.0*cm)
    styles = getSampleStyleSheet()
    BRAND  = colors.HexColor("#4f72fb")

    title_s = ParagraphStyle("T",    parent=styles["Title"],  fontSize=18,
                              textColor=BRAND, spaceAfter=4)
    sub_s   = ParagraphStyle("S",    parent=styles["Normal"], fontSize=10,
                              textColor=colors.grey, spaceAfter=10)
    cell_s  = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=7.5, leading=10)
    hdr_s   = ParagraphStyle("Hdr",  parent=styles["Normal"], fontSize=8,
                              textColor=colors.white, fontName="Helvetica-Bold", leading=10)

    story = [
        Paragraph(
            f"Collegedunia Rankings — "
            f"{source.replace('pipeline','Full Pipeline').replace('quickrun','Quick Run')}",
            title_s
        ),
        Paragraph(f"Exported {datetime.now().strftime('%Y-%m-%d %H:%M')}  •  {len(rows)-1} colleges", sub_s),
        HRFlowable(width="100%", thickness=1.5, color=BRAND),
        Spacer(1, 0.3*cm),
    ]

    headers_raw = rows[0]
    data_raw    = rows[1:]

    # Detect silo columns using proper header matching (ADMISSIONS, FEES, etc.)
    silo_rank_cols, silo_url_cols, updated_cols = _detect_silo_cols(headers_raw)

    # Only include silos that actually appear in this sheet
    present_silos = [st for st in SILO_TYPES if st in silo_rank_cols]

    # ── Column widths — fill full usable page width ──────────────────
    page_w    = landscape(A4)[0] - 1.6*cm   # usable ≈ 27cm (0.8cm margins)
    id_w      = 1.8*cm
    name_w    = 7.0*cm
    course_w  = 4.0*cm
    upd_w     = 2.8*cm
    fixed_w   = id_w + name_w + course_w + upd_w
    n_silos   = max(len(present_silos), 1)
    silo_unit = (page_w - fixed_w) / n_silos

    col_widths = [id_w, name_w, course_w] + [silo_unit] * len(present_silos) + [upd_w]

    # ── Header row ───────────────────────────────────────────────────
    hdr_row = [
        Paragraph("College ID",   hdr_s),
        Paragraph("College Name", hdr_s),
        Paragraph("Course",       hdr_s),
    ]
    for st in present_silos:
        hdr_row.append(Paragraph(SILO_DISPLAY.get(st, st), hdr_s))
    hdr_row.append(Paragraph("Last Updated", hdr_s))

    table_data = [hdr_row]

    # ── Data rows ────────────────────────────────────────────────────
    # Horizontal padding applied to every cell (LEFTPADDING + RIGHTPADDING).
    CELL_PAD = 8
    for row in data_raw:
        cid    = str(row[0]).strip() if row else ""
        cname  = str(row[2]).strip() if len(row) > 2 else ""
        course = str(row[3]).strip() if len(row) > 3 else ""

        # Latest valid rank for each silo (most recent run column first)
        silo_vals = []
        for st in present_silos:
            cols = silo_rank_cols.get(st, [])
            val  = "—"
            for ci in reversed(cols):
                if len(row) > ci:
                    v = str(row[ci]).strip()
                    if v:
                        val = v
                        break
            silo_vals.append(val)

        # Last updated timestamp
        updated = ""
        for uc in reversed(updated_cols):
            if len(row) > uc and str(row[uc]).strip():
                updated = str(row[uc]).strip()[:16]
                break

        # Cell order must match col_widths so each value is clipped to its column.
        raw_cells = [cid, cname, course] + silo_vals + [updated]
        data_row = [
            Paragraph(
                _ep(_truncate_to_width(val, col_widths[i] - CELL_PAD, "Helvetica", 7.5)),
                cell_s
            )
            for i, val in enumerate(raw_cells)
        ]
        table_data.append(data_row)

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  BRAND),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 7.5),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.HexColor("#f5f7ff"), colors.white]),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#d0d4e8")),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(t)

    doc.build(story)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Reports
# ══════════════════════════════════════════════════════════════════════════════

# ── Silo type constants ───────────────────────────────────────────────────────
SILO_TYPES = ['ADMISSIONS', 'FEES', 'PLACEMENTS', 'SCHOLARSHIPS', 'MAIN', 'SINGLE_COURSE']
SILO_DISPLAY = {
    'ADMISSIONS':    'Admissions',
    'FEES':          'Fees',
    'PLACEMENTS':    'Placements',
    'SCHOLARSHIPS':  'Scholarships',
    'MAIN':          'Main',
    'SINGLE_COURSE': 'Single Course',
}


def _detect_silo_cols(headers):
    """
    From sheet headers return:
      silo_rank_cols : { silo_type -> [col_index, ...] }   (rank/value columns)
      silo_url_cols  : { silo_type -> [col_index, ...] }   (URL columns)
      updated_cols   : [col_index, ...]
    Works for both  ADMISSIONS (run-1 no suffix)
    and             ADMISSIONS_R2  (run-N with suffix).
    """
    silo_rank_cols = {}
    silo_url_cols  = {}
    updated_cols   = []

    for i, h in enumerate(headers):
        if i <= 4:
            continue
        h_u = h.upper().replace(' ', '_')
        if 'UPDATED' in h_u:
            updated_cols.append(i)
            continue
        is_url = '_URL' in h_u
        for st in SILO_TYPES:
            if h_u == st \
               or h_u == st + '_URL' \
               or h_u.startswith(st + '_R') \
               or h_u.startswith(st + '_URL_R'):
                if is_url:
                    silo_url_cols.setdefault(st, []).append(i)
                else:
                    silo_rank_cols.setdefault(st, []).append(i)
                break

    return silo_rank_cols, silo_url_cols, updated_cols


def _build_report_data(period: str = "daily") -> dict:
    today     = date.today()
    since     = today if period == "daily" else today - timedelta(days=6)
    since_str = since.isoformat()

    state   = load_state()
    rows    = get_sheet_data("Final",           max_rows=5000)
    qr_rows = get_sheet_data("Quick Run Final", max_rows=5000)

    base = {
        "period":        period,
        "generated_at":  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "run_number":    state.get("run_number", "—"),
        "last_run_date": state.get("phase1_last_run", "Never"),
        "since_date":    since_str,
    }

    has_pipeline = len(rows) >= 2
    has_qr       = len(qr_rows) >= 2

    # ── FULL PIPELINE ────────────────────────────────────────────────────────
    pipeline_rows    = []
    total_keywords   = 0
    total_colleges   = 0
    ranking_gain     = 0
    ranking_dropped  = 0
    tracked_on       = state.get("phase1_last_run", "")
    present_silos    = []
    silo_type_counts = {}
    ranked_count     = 0
    not_found_count  = 0

    if has_pipeline:
        headers        = rows[0]
        data           = rows[1:]
        total_keywords = len(data)
        total_colleges = len({str(r[0]).strip() for r in data if r})

        silo_rank_cols, _, updated_cols = _detect_silo_cols(headers)
        present_silos = [st for st in SILO_TYPES if st in silo_rank_cols]

        for row in data:
            college_id = str(row[0]).strip() if row else ""
            keywords   = str(row[4]).strip() if len(row) > 4 else ""
            silo_data  = {}
            row_ranked = False
            row_nf     = False

            for st in present_silos:
                cols   = silo_rank_cols[st]
                n_runs = len(cols)

                # Latest run values
                latest_rank = str(row[cols[-1]]).strip() if len(row) > cols[-1] else ""
                latest_date = ""
                if updated_cols:
                    uc = updated_cols[-1]
                    if len(row) > uc:
                        latest_date = str(row[uc]).strip()[:10]

                # Previous run values (only if 2+ runs exist)
                prev_rank = ""
                prev_date = ""
                if n_runs >= 2:
                    prev_rank = str(row[cols[-2]]).strip() if len(row) > cols[-2] else ""
                    if len(updated_cols) >= 2:
                        uc2 = updated_cols[-2]
                        if len(row) > uc2:
                            prev_date = str(row[uc2]).strip()[:10]

                silo_data[st] = {
                    "prev_rank":   prev_rank,
                    "prev_date":   prev_date,
                    "latest_rank": latest_rank,
                    "latest_date": latest_date,
                }

                # Stats: silo counts, ranked/not-found, gain/drop
                lr_val = latest_rank
                pr_val = prev_rank
                if lr_val and lr_val not in ("NOT_FOUND", "ERROR", "CLEARED"):
                    row_ranked = True
                    disp = SILO_DISPLAY.get(st, st)
                    silo_type_counts[disp] = silo_type_counts.get(disp, 0) + 1
                    try:
                        lr_num = int(lr_val)
                        if not pr_val or pr_val in ("NOT_FOUND", "ERROR", "CLEARED"):
                            ranking_gain += 1          # newly ranked
                        else:
                            try:
                                if lr_num < int(pr_val):
                                    ranking_gain += 1  # rank improved
                            except ValueError:
                                pass
                    except ValueError:
                        pass
                elif lr_val == "NOT_FOUND":
                    row_nf = True
                    if pr_val and pr_val not in ("NOT_FOUND", "ERROR", "CLEARED"):
                        try:
                            int(pr_val)
                            ranking_dropped += 1       # was ranked, now dropped
                        except ValueError:
                            pass

            if row_ranked:
                ranked_count += 1
            elif row_nf:
                not_found_count += 1

            pipeline_rows.append({
                "college_id": college_id,
                "keywords":   keywords,
                "silos":      silo_data,
            })

    # ── QUICK RUN ────────────────────────────────────────────────────────────
    qr_total_keywords = 0
    qr_total_colleges = 0
    qr_tracked_on     = ""
    qr_batch          = 0
    qr_single         = 0

    if has_qr:
        qr_data           = qr_rows[1:]
        qr_total_keywords = len(qr_data)
        qr_total_colleges = len({str(r[0]).strip() for r in qr_data if r})
        if qr_total_colleges > 1:
            qr_batch  = 1
        elif qr_total_colleges == 1:
            qr_single = 1

        _, _, qr_uc = _detect_silo_cols(qr_rows[0])
        for qr_row in qr_data:
            for uc in reversed(qr_uc):
                if len(qr_row) > uc and str(qr_row[uc]).strip():
                    ts = str(qr_row[uc]).strip()[:10]
                    if not qr_tracked_on or ts > qr_tracked_on:
                        qr_tracked_on = ts
                    break

    base.update({
        # Full Pipeline
        "has_pipeline":    has_pipeline,
        "total_colleges":  total_colleges,
        "total_keywords":  total_keywords,
        "tracked_on":      tracked_on,
        "ranking_gain":    ranking_gain,
        "ranking_dropped": ranking_dropped,
        "present_silos":   present_silos,
        "pipeline_rows":   pipeline_rows,
        # Quick Run
        "has_qr":             has_qr,
        "qr_total_colleges":  qr_total_colleges,
        "qr_total_keywords":  qr_total_keywords,
        "qr_tracked_on":      qr_tracked_on,
        "qr_batch":           qr_batch,
        "qr_single":          qr_single,
        # Legacy / chart fields (kept for backward compat)
        "ranked_count":    ranked_count,
        "not_found":       not_found_count,
        "error_count":     max(0, total_keywords - ranked_count - not_found_count),
        "silo_breakdown":  silo_type_counts,
        "period_colleges": [],
        "period_count":    0,
        "qr_total":        qr_total_keywords,
        "qr_last_updated": qr_tracked_on,
        "qr_colleges":     [],
    })
    return base


@app.route("/api/report/daily")
def api_report_daily():
    return jsonify(_build_report_data("daily"))


@app.route("/api/report/weekly")
def api_report_weekly():
    return jsonify(_build_report_data("weekly"))


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Download Report PDF
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/download/pdf")
def api_download_pdf():
    period = request.args.get("period", "daily")
    try:
        report  = _build_report_data(period)
        pdf_buf = _generate_report_pdf(report)
        fname   = f"ranking_report_{period}_{date.today().isoformat()}.pdf"
        return send_file(pdf_buf, mimetype="application/pdf",
                         as_attachment=True, download_name=fname)
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        log.exception("Report PDF error")
        return jsonify({"error": str(e)}), 500


def _generate_report_pdf(report: dict) -> BytesIO:
    try:
        from reportlab.lib.pagesizes  import A4, landscape
        from reportlab.lib.styles     import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units      import cm
        from reportlab.lib            import colors
        from reportlab.platypus       import (
            SimpleDocTemplate, Paragraph, Spacer,
            Table, TableStyle, HRFlowable
        )
    except ImportError:
        raise RuntimeError("reportlab not installed — run: pip3 install reportlab --break-system-packages")

    buf  = BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             leftMargin=0.8*cm, rightMargin=0.8*cm,
                             topMargin=1.2*cm,  bottomMargin=1.2*cm)
    stys  = getSampleStyleSheet()
    BRAND = colors.HexColor("#4f72fb")
    WARN  = colors.HexColor("#f5a623")

    def _style(name, parent="Normal", **kw):
        return ParagraphStyle(name, parent=stys[parent], **kw)

    title_s = _style("RTitle", "Title",   fontSize=18, textColor=BRAND, spaceAfter=4)
    sub_s   = _style("RSub",              fontSize=10, textColor=colors.grey, spaceAfter=10)
    h2_s    = _style("RH2",   "Heading2", fontSize=12, textColor=BRAND,
                     spaceBefore=12, spaceAfter=5)
    h3_s    = _style("RH3",   "Heading3", fontSize=10, textColor=BRAND,
                     spaceBefore=8,  spaceAfter=3)
    warn_s  = _style("RWarn",             fontSize=11, textColor=WARN)

    story = [
        Paragraph("Collegedunia Article Ranking Pipeline", title_s),
        Paragraph(
            f"{report['period'].capitalize()} Report  •  Generated {report['generated_at']}",
            sub_s
        ),
        HRFlowable(width="100%", thickness=1.5, color=BRAND),
        Spacer(1, 0.4*cm),
    ]

    # ── Pipeline Summary ──────────────────────────────────────────────────────
    story.append(Paragraph("Pipeline Summary", h2_s))
    if report.get("has_pipeline"):
        summary = [
            ["Metric", "Value"],
            ["Total Colleges",         str(report.get("total_colleges",  0))],
            ["Total Keywords",          str(report.get("total_keywords",  0))],
            ["Tracked On",              str(report.get("tracked_on",      "—"))],
            ["Ranking Gain",            str(report.get("ranking_gain",    0))],
            ["Ranking Dropped",         str(report.get("ranking_dropped", 0))],
            ["Report Generation Date",  str(report.get("generated_at",   "—"))],
        ]
        _table(story, summary, [10*cm, 8*cm], BRAND)
    else:
        story.append(Paragraph("no pipeline_run running", warn_s))
    story.append(Spacer(1, 0.4*cm))

    # ── Full Pipeline Data (one table per silo) ───────────────────────────────
    silos = report.get("present_silos", [])
    prows = report.get("pipeline_rows", [])
    if report.get("has_pipeline") and silos and prows:
        story.append(Paragraph("Full Pipeline Data", h2_s))
        _pw    = landscape(A4)[0] - 1.6*cm   # ≈ 27cm usable
        id_w   = 1.8*cm;  kw_w = 8.0*cm
        remaining = (_pw - id_w - kw_w) / 4.0   # 4 rank/date cols
        rank_w = remaining;  dt_w = remaining
        col_w  = [id_w, kw_w, rank_w, dt_w, rank_w, dt_w]

        for st in silos:
            disp = SILO_DISPLAY.get(st, st)
            story.append(Paragraph(disp, h3_s))
            hdr   = ["College ID", "Keywords",
                     "Prev Rank", "Prev Date", "Latest Rank", "Latest Date"]
            tdata = [hdr]
            for row in prows:
                sd = row["silos"].get(st, {})
                tdata.append([
                    row["college_id"],
                    row["keywords"],
                    sd.get("prev_rank",   ""),
                    sd.get("prev_date",   ""),
                    sd.get("latest_rank", ""),
                    sd.get("latest_date", ""),
                ])
            _table(story, tdata, col_w, BRAND)
            story.append(Spacer(1, 0.3*cm))

    # ── Quick Run Summary ─────────────────────────────────────────────────────
    story.append(Paragraph("Quick Run Summary", h2_s))
    if report.get("has_qr"):
        qr_sum = [
            ["Metric", "Value"],
            ["Total Colleges",         str(report.get("qr_total_colleges", 0))],
            ["Total Keywords",          str(report.get("qr_total_keywords", 0))],
            ["Tracked On",              str(report.get("qr_tracked_on",     "—"))],
            ["Report Generation Date",  str(report.get("generated_at",      "—"))],
        ]
        _table(story, qr_sum, [10*cm, 8*cm], BRAND)
    else:
        story.append(Paragraph("no quick_run run", warn_s))

    doc.build(story)
    buf.seek(0)
    return buf


def _table(story, data, col_widths, brand_color):
    from reportlab.lib            import colors
    from reportlab.lib.styles     import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus       import Table, TableStyle, Paragraph

    stys   = getSampleStyleSheet()
    cell_s = ParagraphStyle("TC", parent=stys["Normal"], fontSize=9,  leading=12)
    hdr_s  = ParagraphStyle("TH", parent=stys["Normal"], fontSize=9,  leading=12,
                             textColor=colors.white, fontName="Helvetica-Bold")

    def _ep(s):
        return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    # Clip each cell to its column so text stays inside the borders (PDF only).
    CELL_PAD = 14   # LEFTPADDING + RIGHTPADDING
    wrapped = []
    for ri, row in enumerate(data):
        is_hdr = ri == 0
        fname  = "Helvetica-Bold" if is_hdr else "Helvetica"
        wrapped.append([
            Paragraph(
                _ep(_truncate_to_width(cell, col_widths[ci] - CELL_PAD, fname, 9)),
                hdr_s if is_hdr else cell_s
            )
            for ci, cell in enumerate(row)
        ])

    t = Table(wrapped, colWidths=col_widths)
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  brand_color),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.HexColor("#f5f7ff"), colors.white]),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#d0d4e8")),
        ("LEFTPADDING",   (0, 0), (-1, -1), 7),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 7),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(t)


def _generate_report_csv(report: dict) -> BytesIO:
    """CSV report: pipeline summary + full pipeline data + QR summary."""
    buf = StringIO()
    w   = csv.writer(buf)

    # ── Pipeline Summary ──────────────────────────────────────────────────────
    w.writerow(["=== PIPELINE SUMMARY ==="])
    if report.get("has_pipeline"):
        w.writerow(["Metric", "Value"])
        w.writerow(["Total Colleges",         report.get("total_colleges",  0)])
        w.writerow(["Total Keywords",          report.get("total_keywords",  0)])
        w.writerow(["Tracked On",              report.get("tracked_on",      "")])
        w.writerow(["Ranking Gain",            report.get("ranking_gain",    0)])
        w.writerow(["Ranking Dropped",         report.get("ranking_dropped", 0)])
        w.writerow(["Report Generation Date",  report.get("generated_at",   "")])
    else:
        w.writerow(["Status", "no pipeline_run running"])
    w.writerow([])

    # ── Full Pipeline Data ────────────────────────────────────────────────────
    w.writerow(["=== FULL PIPELINE DATA ==="])
    silos = report.get("present_silos", [])
    if report.get("has_pipeline") and silos:
        h1 = ["College_Id", "Keywords"]
        for st in silos:
            h1 += [SILO_DISPLAY.get(st, st), "", "", ""]
        h2 = ["", ""]
        for _ in silos:
            h2 += ["Previous Rank", "Previous Rank Date", "Latest Rank", "Latest Rank Date"]
        w.writerow(h1)
        w.writerow(h2)
        for row in report.get("pipeline_rows", []):
            data_row = [row["college_id"], row["keywords"]]
            for st in silos:
                sd = row["silos"].get(st, {})
                data_row += [
                    sd.get("prev_rank",   ""),
                    sd.get("prev_date",   ""),
                    sd.get("latest_rank", ""),
                    sd.get("latest_date", ""),
                ]
            w.writerow(data_row)
    else:
        w.writerow(["no pipeline_run running"])
    w.writerow([])

    # ── Quick Run Summary ─────────────────────────────────────────────────────
    w.writerow(["=== QUICK RUN SUMMARY ==="])
    if report.get("has_qr"):
        w.writerow(["Metric", "Value"])
        w.writerow(["Total Colleges",         report.get("qr_total_colleges", 0)])
        w.writerow(["Total Keywords",          report.get("qr_total_keywords", 0)])
        w.writerow(["Tracked On",              report.get("qr_tracked_on",     "")])
        w.writerow(["Report Generation Date",  report.get("generated_at",      "")])
    else:
        w.writerow(["Status", "no quick_run run"])

    buf.seek(0)
    return BytesIO(buf.getvalue().encode("utf-8"))


@app.route("/api/download/report-csv")
def api_download_report_csv():
    period = request.args.get("period", "daily")
    try:
        report  = _build_report_data(period)
        csv_buf = _generate_report_csv(report)
        fname   = f"ranking_report_{period}_{date.today().isoformat()}.csv"
        return send_file(csv_buf, mimetype="text/csv",
                         as_attachment=True, download_name=fname)
    except Exception as e:
        log.exception("Report CSV error")
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Comparison XLSX  (colored Excel download)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/compare/xlsx", methods=["POST"])
def api_compare_xlsx():
    body = request.get_json(force=True, silent=True) or {}
    try:
        buf   = _generate_comparison_xlsx(body)
        fname = f"comparison_{date.today().isoformat()}.xlsx"
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name=fname
        )
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        log.exception("Comparison XLSX error")
        return jsonify({"error": str(e)}), 500


def _generate_comparison_xlsx(data: dict) -> BytesIO:
    try:
        import openpyxl
        from openpyxl.styles   import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils    import get_column_letter
    except ImportError:
        raise RuntimeError(
            "openpyxl not installed — run: pip3 install openpyxl --break-system-packages"
        )

    silos    = data.get("present_silos", [])
    rows     = data.get("rows",          [])
    old_date = data.get("old_date",      "")
    new_date = data.get("new_date",      "")

    prev_lbl   = f"Prev Rank{' (' + old_date + ')' if old_date else ''}"
    latest_lbl = f"Latest Rank{' (' + new_date + ')' if new_date else ''}"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison"

    # ── Fills ────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="4F72FB")   # brand blue  — header row
    SUB_FILL  = PatternFill("solid", fgColor="D1D5E8")   # light slate — sub-header row
    GREEN_FILL = PatternFill("solid", fgColor="D1FAE5")  # improved
    RED_FILL   = PatternFill("solid", fgColor="FEE2E2")  # dropped
    YELL_FILL  = PatternFill("solid", fgColor="FEF9C3")  # same / N/A

    WHITE_BOLD = Font(bold=True, color="FFFFFF")
    BOLD       = Font(bold=True)
    CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin = Side(style="thin", color="D0D4E8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: silo group headers ─────────────────────────────
    # Columns: College_Id(1), Keywords(2), then pairs per silo
    BASE_COLS = 2
    h1_vals = ["College_Id", "Keywords"]
    for st in silos:
        h1_vals.append(SILO_DISPLAY.get(st, st))
        h1_vals.append("")   # merged with next
    ws.append(h1_vals)

    # ── Row 2: sub-column headers ─────────────────────────────
    h2_vals = ["", ""]
    for _ in silos:
        h2_vals.extend([prev_lbl, latest_lbl])
    ws.append(h2_vals)

    # Style row 1
    for c in range(1, len(h1_vals) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font      = WHITE_BOLD
        cell.fill      = HDR_FILL
        cell.alignment = CENTER
        cell.border    = border

    # Merge silo name across 2 cols in row 1
    for si in range(len(silos)):
        c1 = BASE_COLS + 1 + si * 2
        c2 = c1 + 1
        ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)

    # Style row 2
    for c in range(1, len(h2_vals) + 1):
        cell = ws.cell(row=2, column=c)
        cell.font      = BOLD
        cell.fill      = SUB_FILL
        cell.alignment = CENTER
        cell.border    = border

    # ── Data rows ─────────────────────────────────────────────
    for row_data in rows:
        dr = [
            row_data.get("college_id", ""),
            row_data.get("keywords",   ""),
        ]
        for st in silos:
            sd = (row_data.get("silos") or {}).get(st, {})
            dr.append(sd.get("prev_rank",   "") or "")
            dr.append(sd.get("latest_rank", "") or "")
        ws.append(dr)

        row_idx = ws.max_row
        # Style base columns
        for c in range(1, BASE_COLS + 1):
            ws.cell(row=row_idx, column=c).alignment = LEFT
            ws.cell(row=row_idx, column=c).border    = border

        # Colour Latest Rank cells
        for si, st in enumerate(silos):
            pr_col = BASE_COLS + 1 + si * 2       # Previous Rank col
            lr_col = pr_col + 1                    # Latest Rank col

            pr_cell = ws.cell(row=row_idx, column=pr_col)
            lr_cell = ws.cell(row=row_idx, column=lr_col)

            pr_cell.alignment = CENTER
            lr_cell.alignment = CENTER
            pr_cell.border    = border
            lr_cell.border    = border

            lr = str(lr_cell.value or "").strip()
            pr = str(pr_cell.value or "").strip()

            fill = YELL_FILL
            if lr and lr not in ("NOT_FOUND", "ERROR", ""):
                try:
                    lr_n = int(lr)
                    pr_n = int(pr)
                    if   lr_n < pr_n: fill = GREEN_FILL
                    elif lr_n > pr_n: fill = RED_FILL
                    else:             fill = YELL_FILL
                except (ValueError, TypeError):
                    fill = YELL_FILL
            if lr:
                lr_cell.fill = fill

    # ── Column widths (auto-fit, capped) ─────────────────────
    col_width_hints = {1: 12, 2: 30}   # College_Id, Keywords
    for si in range(len(silos)):
        col_width_hints[BASE_COLS + 1 + si * 2]     = 14  # Prev Rank
        col_width_hints[BASE_COLS + 1 + si * 2 + 1] = 14  # Latest Rank

    for col_idx, hint_w in col_width_hints.items():
        # scan actual content width
        max_len = hint_w
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v:
                max_len = max(max_len, min(len(str(v)) + 2, 35))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A3"   # freeze header rows

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Comparison PDF
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/compare/pdf", methods=["POST"])
def api_compare_pdf():
    body = request.get_json(force=True, silent=True) or {}
    try:
        pdf_buf = _generate_comparison_pdf(body)
        fname   = f"comparison_{date.today().isoformat()}.pdf"
        return send_file(pdf_buf, mimetype="application/pdf",
                         as_attachment=True, download_name=fname)
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        log.exception("Comparison PDF error")
        return jsonify({"error": str(e)}), 500


def _generate_comparison_pdf(data: dict) -> BytesIO:
    try:
        from reportlab.lib.pagesizes  import A4, landscape
        from reportlab.lib.styles     import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units      import cm
        from reportlab.lib            import colors
        from reportlab.platypus       import (
            SimpleDocTemplate, Paragraph, Spacer,
            Table, TableStyle, HRFlowable
        )
    except ImportError:
        raise RuntimeError("reportlab not installed — run: pip3 install reportlab --break-system-packages")

    def _ep(s):
        return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    buf  = BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             leftMargin=0.8*cm, rightMargin=0.8*cm,
                             topMargin=1.2*cm,  bottomMargin=1.2*cm)
    base_stys = getSampleStyleSheet()
    BRAND   = colors.HexColor("#4f72fb")
    C_GREEN = colors.HexColor("#d1fae5")   # light green — improved rank
    C_RED   = colors.HexColor("#fee2e2")   # light red   — worsened rank
    C_YELL  = colors.HexColor("#fef9c3")   # light yellow — same / N/A

    title_s = ParagraphStyle("CTitle", parent=base_stys["Title"],    fontSize=18,
                              textColor=BRAND, spaceAfter=4)
    sub_s   = ParagraphStyle("CSub",   parent=base_stys["Normal"],   fontSize=10,
                              textColor=colors.grey, spaceAfter=10)
    h2_s    = ParagraphStyle("CH2",    parent=base_stys["Heading2"], fontSize=12,
                              textColor=BRAND, spaceBefore=12, spaceAfter=5)
    h3_s    = ParagraphStyle("CH3",    parent=base_stys["Heading3"], fontSize=10,
                              textColor=BRAND, spaceBefore=8,  spaceAfter=3)
    warn_s  = ParagraphStyle("CWarn",  parent=base_stys["Normal"],   fontSize=11,
                              textColor=colors.HexColor("#f5a623"))
    cell_s  = ParagraphStyle("CCell",  parent=base_stys["Normal"],   fontSize=9, leading=12)
    hdr_s   = ParagraphStyle("CHdr",   parent=base_stys["Normal"],   fontSize=9, leading=12,
                              textColor=colors.white, fontName="Helvetica-Bold")

    story = [
        Paragraph("Collegedunia Ranking Comparison", title_s),
        Paragraph(f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", sub_s),
        HRFlowable(width="100%", thickness=1.5, color=BRAND),
        Spacer(1, 0.4*cm),
    ]

    silos    = data.get("present_silos", [])
    rows     = data.get("rows", [])
    old_date = data.get("old_date", "")
    new_date = data.get("new_date", "")

    prev_lbl   = f"Prev Rank{' (' + old_date + ')' if old_date else ''}"
    latest_lbl = f"Latest Rank{' (' + new_date + ')' if new_date else ''}"

    if not silos or not rows:
        story.append(Paragraph("No comparison data available.", warn_s))
    else:
        story.append(Paragraph(f"Comparison Data — {len(rows)} rows", h2_s))
        page_w   = landscape(A4)[0] - 1.6*cm        # usable width ≈ 27cm
        id_w2    = 2.0*cm
        kw_w2    = 8.5*cm
        remaining= (page_w - id_w2 - kw_w2) / 4.0  # 4 equal rank/date cols ≈ 4.1cm
        col_w    = [id_w2, kw_w2, remaining, remaining, remaining, remaining]
        HDR      = ["College ID", "Keywords", prev_lbl, "Prev Date", latest_lbl, "Latest Date"]
        CELL_PAD = 14

        for st in silos:
            disp = SILO_DISPLAY.get(st, st)
            story.append(Paragraph(disp, h3_s))

            hdr_row = [
                Paragraph(
                    _ep(_truncate_to_width(h, col_w[ci] - CELL_PAD, "Helvetica-Bold", 9)),
                    hdr_s
                )
                for ci, h in enumerate(HDR)
            ]
            tdata    = [hdr_row]
            bg_cmds  = []   # (row_idx_1based, bg_color) for latest_rank col (index 4)

            for ri, row in enumerate(rows, start=1):
                sd   = (row.get("silos") or {}).get(st, {})
                pr   = str(sd.get("prev_rank",   "") or "")
                pd_v = str(sd.get("prev_date",   "") or "")
                lr   = str(sd.get("latest_rank", "") or "")
                ld   = str(sd.get("latest_date", "") or "")

                # Determine background colour for the latest_rank cell
                bg = C_YELL
                try:
                    lr_num = int(lr)
                    try:
                        pr_num = int(pr)
                        if   lr_num < pr_num: bg = C_GREEN   # improved (lower = better)
                        elif lr_num > pr_num: bg = C_RED     # worsened
                        else:                 bg = C_YELL    # identical
                    except ValueError:
                        bg = C_YELL
                except ValueError:
                    bg = C_YELL

                bg_cmds.append((ri, bg))

                vals = [
                    row.get("college_id", ""), row.get("keywords", ""),
                    pr, pd_v, lr, ld
                ]
                tdata.append([
                    Paragraph(
                        _ep(_truncate_to_width(str(v), col_w[ci] - CELL_PAD, "Helvetica", 9)),
                        cell_s
                    )
                    for ci, v in enumerate(vals)
                ])

            ts_cmds = [
                ("BACKGROUND",    (0, 0), (-1, 0),  BRAND),
                ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
                ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
                ("FONTSIZE",      (0, 0), (-1, -1), 9),
                ("ROWBACKGROUNDS",(0, 1), (-1, -1),
                 [colors.HexColor("#f5f7ff"), colors.white]),
                ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#d0d4e8")),
                ("LEFTPADDING",   (0, 0), (-1, -1), 7),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 7),
                ("TOPPADDING",    (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("VALIGN",        (0, 0), (-1, -1), "TOP"),
            ]
            # Per-cell colour overrides — appended AFTER ROWBACKGROUNDS so they win
            for (ri, bg) in bg_cmds:
                ts_cmds.append(("BACKGROUND", (4, ri), (4, ri), bg))

            t = Table(tdata, colWidths=col_w)
            t.setStyle(TableStyle(ts_cmds))
            story.append(t)
            story.append(Spacer(1, 0.3*cm))

    doc.build(story)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Email
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/email/config", methods=["GET", "POST"])
def api_email_config():
    cfg = load_ui_config()
    if request.method == "GET":
        safe = {k: v for k, v in cfg.items() if k != "smtp_password"}
        safe["smtp_password_set"] = bool(cfg.get("smtp_password"))
        return jsonify(safe)

    body = request.get_json(force=True, silent=True) or {}
    for key in ("smtp_host", "smtp_port", "smtp_user", "smtp_from"):
        if key in body:
            cfg[key] = body[key]
    new_pw = body.get("smtp_password", "")
    # Only update password when user typed something real
    # (not empty, not the masked bullet placeholder the UI sends when unchanged)
    if new_pw and set(new_pw) != {'•'}:
        cfg["smtp_password"] = new_pw
    try:
        cfg["smtp_port"] = int(cfg.get("smtp_port", 587))
    except Exception:
        cfg["smtp_port"] = 587
    save_ui_config(cfg)
    return jsonify({"message": "SMTP configuration saved."})


@app.route("/api/email/recipients", methods=["GET", "POST"])
def api_email_recipients():
    cfg = load_ui_config()
    if request.method == "GET":
        return jsonify({"recipients": cfg.get("recipients", [])})

    body  = request.get_json(force=True, silent=True) or {}
    email = (body.get("email") or "").strip().lower()
    name  = (body.get("name") or "").strip()

    if not email or "@" not in email:
        return jsonify({"error": "Valid email address required."}), 400

    recipients = cfg.get("recipients", [])
    if any(r["email"] == email for r in recipients):
        return jsonify({"error": "Recipient already exists."}), 409

    recipients.append({"name": name, "email": email})
    cfg["recipients"] = recipients
    save_ui_config(cfg)
    return jsonify({"message": f"Added {email}.", "recipients": recipients})


@app.route("/api/email/recipients/<path:email>", methods=["DELETE"])
def api_delete_recipient(email):
    cfg        = load_ui_config()
    recipients = cfg.get("recipients", [])
    original   = len(recipients)
    recipients = [r for r in recipients if r["email"].lower() != email.lower()]
    if len(recipients) == original:
        return jsonify({"error": "Recipient not found."}), 404
    cfg["recipients"] = recipients
    save_ui_config(cfg)
    return jsonify({"message": f"Removed {email}.", "recipients": recipients})


@app.route("/api/email/send", methods=["POST"])
def api_send_email():
    body   = request.get_json(force=True, silent=True) or {}
    period = body.get("period", "daily")
    fmt    = body.get("format", "pdf").lower()   # "pdf" or "csv"

    cfg = load_ui_config()
    if not cfg.get("smtp_host"):
        return jsonify({"error": "SMTP not configured. Set it in the Email tab."}), 400

    recipients = cfg.get("recipients", [])
    if not recipients:
        return jsonify({"error": "No recipients. Add at least one in the Email tab."}), 400

    try:
        report = _build_report_data(period)
        if fmt == "csv":
            file_buf   = _generate_report_csv(report)
            file_bytes = file_buf.read()
            mime_main  = "text"; mime_sub = "csv"
            fname      = f"ranking_report_{period}_{date.today().isoformat()}.csv"
            fmt_label  = "CSV"
        else:
            file_buf   = _generate_report_pdf(report)
            file_bytes = file_buf.read()
            mime_main  = "application"; mime_sub = "pdf"
            fname      = f"ranking_report_{period}_{date.today().isoformat()}.pdf"
            fmt_label  = "PDF"
    except Exception as e:
        return jsonify({"error": f"Could not generate {fmt.upper()}: {e}"}), 500

    subject = f"Collegedunia Ranking Report — {period.capitalize()} ({date.today().isoformat()})"

    body_html = f"""
    <html><body style="font-family:sans-serif;color:#333;max-width:600px;">
    <div style="background:#4f72fb;padding:20px;border-radius:10px 10px 0 0;">
      <h2 style="color:white;margin:0;">Collegedunia Ranking Pipeline</h2>
      <p style="color:#c7d2fe;margin:4px 0 0;">{period.capitalize()} Report ({fmt_label})</p>
    </div>
    <div style="background:#f8f9ff;padding:20px;border:1px solid #e0e4ff;border-top:none;border-radius:0 0 10px 10px;">
      <table cellpadding="8" cellspacing="0" style="width:100%;border-collapse:collapse;">
        <tr style="background:#4f72fb;color:white;"><th>Metric</th><th>Value</th></tr>
        <tr style="background:#f0f2ff;"><td>Generated At</td><td>{report['generated_at']}</td></tr>
        <tr><td>Run Number</td><td>{report.get('run_number','—')}</td></tr>
        <tr style="background:#f0f2ff;"><td>Total Colleges</td><td>{report.get('total_colleges', 0)}</td></tr>
        <tr><td>Total Keywords</td><td>{report.get('total_keywords', 0)}</td></tr>
        <tr style="background:#f0f2ff;"><td>Ranking Gain</td><td><b style="color:#10b981;">{report.get('ranking_gain', 0)}</b></td></tr>
        <tr><td>Ranking Dropped</td><td>{report.get('ranking_dropped', 0)}</td></tr>
        <tr style="background:#f0f2ff;"><td>QR Total Keywords</td><td>{report.get('qr_total_keywords', 0)}</td></tr>
      </table>
      <p style="margin-top:16px;font-size:12px;color:#888;">
        Full details are in the attached {fmt_label}. Sent by Collegedunia Ranking Pipeline UI.
      </p>
    </div>
    </body></html>
    """

    errors = []
    for rec in recipients:
        try:
            msg = MIMEMultipart("mixed")
            msg["Subject"] = subject
            msg["From"]    = cfg.get("smtp_from") or cfg.get("smtp_user")
            msg["To"]      = rec["email"]
            msg.attach(MIMEText(body_html, "html"))
            part = MIMEBase(mime_main, mime_sub)
            part.set_payload(file_bytes)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f'attachment; filename="{fname}"')
            msg.attach(part)
            with smtplib.SMTP(cfg["smtp_host"], int(cfg.get("smtp_port", 587))) as srv:
                srv.ehlo(); srv.starttls()
                srv.login(cfg["smtp_user"], cfg["smtp_password"])
                srv.sendmail(msg["From"], rec["email"], msg.as_string())
            log.info(f"Email sent ({fmt_label}) -> {rec['email']}")
        except Exception as e:
            errors.append(f"{rec['email']}: {e}")
            log.warning(f"Email failed -> {rec['email']}: {e}")

    if errors:
        # Surface the first real error so the UI can show it
        first_err = errors[0] if errors else "Unknown error"
        return jsonify({
            "error":   first_err,
            "details": errors,
        }), 500
    return jsonify({"message": f"Report ({fmt_label}) sent to {len(recipients)} recipient(s)."})


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — Logs
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/logs")
def api_logs():
    lines_count = int(request.args.get("lines", 200))
    try:
        with open(PIPELINE_LOG) as f:
            lines = f.readlines()
        return jsonify({"lines": [l.rstrip() for l in lines[-lines_count:]]})
    except Exception:
        return jsonify({"lines": []})


# ══════════════════════════════════════════════════════════════════════════════
#  MIDNIGHT AUTO-RUN — Full Pipeline
# ══════════════════════════════════════════════════════════════════════════════

def _secs_until_midnight() -> float:
    now = datetime.now()
    nxt = (now + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    return max((nxt - now).total_seconds(), 1.0)


def _midnight_pipeline_run():
    """Fires at midnight: runs run_pipeline.py, then reschedules for next midnight."""
    global _next_auto_run
    log.info("Midnight auto-run: Full Pipeline starting")
    can_run = False
    with _lock:
        if not _running["pipeline"] and not _running["quickrun"]:
            _running["pipeline"] = True
            can_run = True

    if can_run:
        def _worker():
            try:
                cmd = [PYTHON_EXE, os.path.join(BASE_DIR, "run_pipeline.py")]
                env = os.environ.copy()
                env["PYTHONUNBUFFERED"] = "1"
                with open(PIPELINE_LOG, "a") as lf:
                    lf.write(
                        f"\n[{datetime.now().isoformat()}]"
                        f" ===== Midnight Auto-Run (Full Pipeline) START =====\n"
                    )
                proc = subprocess.Popen(
                    cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    cwd=BASE_DIR, env=env, text=True, bufsize=1
                )
                for line in proc.stdout:
                    with open(PIPELINE_LOG, "a") as lf:
                        lf.write(line)
                proc.wait()
                with open(PIPELINE_LOG, "a") as lf:
                    lf.write(
                        f"[{datetime.now().isoformat()}]"
                        f" ===== Midnight Auto-Run END (rc={proc.returncode}) =====\n"
                    )
                log.info(f"Midnight auto-run finished (rc={proc.returncode})")
            except Exception:
                log.exception("Midnight auto-run error")
            finally:
                _running["pipeline"] = False
        threading.Thread(target=_worker, name="midnight-autorun", daemon=True).start()
    else:
        log.warning("Midnight auto-run: skipped — pipeline/quickrun already running")

    # Always reschedule for tomorrow midnight
    _schedule_midnight_run()


def _schedule_midnight_run():
    """(Re)schedule the midnight auto-run timer and update _next_auto_run."""
    global _next_auto_run, _midnight_timer
    if _midnight_timer and _midnight_timer.is_alive():
        _midnight_timer.cancel()
    secs = _secs_until_midnight()
    nxt  = (datetime.now() + timedelta(days=1)).replace(
        hour=0, minute=0, second=0, microsecond=0
    )
    _next_auto_run  = nxt.isoformat()
    _midnight_timer = threading.Timer(secs, _midnight_pipeline_run)
    _midnight_timer.daemon = True
    _midnight_timer.start()
    log.info(f"Full Pipeline midnight auto-run scheduled for {_next_auto_run}")


# Start the scheduler immediately when the module loads
_schedule_midnight_run()


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    print(f"\n  ╔══════════════════════════════════════════════════╗")
    print(f"  ║   Collegedunia Ranking UI — http://localhost:{port}  ║")
    print(f"  ╚══════════════════════════════════════════════════╝\n")
    app.run(host="0.0.0.0", port=port, threaded=True, debug=False)
